# import pandas as pd
# import openpyxl 


# df = pd.read_excel('auto.xlsx',index_col=0)
# print(df)

# df_france = df[df["店舗名"] == '和食 清風']
# print(df_france)

# df_france = df[df["店舗名"] == 'オステリアSeifu']
# print(df_france)

# df_france = df[df["店舗名"] == 'Bar Seifu']
# print(df_france)






# wb = openpyxl.load_workbook('auto.xlsx')
# print(wb.sheetnames)

# sheet_new = wb.create_sheet('Sheet_new')
# print(wb.sheetnames)


# wb.save('auto2.xlsx')


# wb = openpyxl.load_workbook('auto2.xlsx')
# print(wb.sheetnames)

# wb.remove ( wb [ '和食 清風' ] )
# wb.remove ( wb [ 'オステリアSeifu' ] )
# wb.remove ( wb [ 'Bar Seifu' ] )

# wb.save('auto2.xlsx')



# wb = openpyxl.load_workbook('auto2.xlsx')
# ws = wb.worksheets[1]
# ws.title = '和食 清風'
# ws = wb.worksheets[2]
# ws.title = 'オステリアSeifu'
# ws = wb.worksheets[3]
# ws.title = 'Bar Seifu'

# wb.save('auto2.xlsx')


# wb =openpyxl.load_workbook('auto2.xlsx')
# sheet = wb.worksheets[0]
# sheet2 = wb.worksheets[1]
# for i in range(1,721):
#    for j in range(1,6): 
#       copy = sheet.cell(row = i, column = j).value

#       sheet2.cell(row = i, column = j, value = copy)
      





# df_france = df[df["店舗名"] == '和食 清風']
# print(df_france)

# wb.save('auto2.xlsx')

import pandas as pd
import openpyxl 
pf=pd.read_excel('auto2.xlsx')
print(pf)
wa=pf[pf["店舗名"]=="和食 清風"]
print(wa)
# print(pf[pf["店舗名"]=="オステリアSeifu "])
wb=pf[pf["店舗名"]=="Bar Seifu"]
wc=pf[pf["店舗名"]=="オステリアSeifu"]

# print(type(wa))
# print(wb)
# wb=openpyxl.Workbook()
# wb.create_sheet(title='和食 清風')
# wb.create_sheet(title='オステリアseifu')
# wb.create_sheet(title='bar sebifu')
# print(wb.sheetnames)
# wb.save('EXCEL/auto1.xlsx')
with pd.ExcelWriter('auto2.xlsx') as writer:
   wa.to_excel(writer, sheet_name='和食 清風')
   wb.to_excel(writer, sheet_name='Bar Seifu')
   wc.to_excel(writer, sheet_name='オステリアSeifu')
# pf.to_excel(writer, sheet_name='Bar Seifu ')
# print(pf)

